import pandas as pd
import argparse
import logging
from colorama import Fore, Style, init
from openpyxl.styles import Font
from openpyxl import load_workbook
import os

# Initialize colorama
init(autoreset=True)

# Logging setup
logging.basicConfig(
    filename="converter.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# -----------------------------
# CLEANING FUNCTION
# -----------------------------
def clean_data(df):
    print(Fore.CYAN + "\n[+] Cleaning data...")

    # Remove duplicates
    before = len(df)
    df = df.drop_duplicates()
    after = len(df)

    print(Fore.YELLOW + f"[+] Removed {before - after} duplicate rows")

    # Fill missing values
    df = df.fillna("N/A")

    # Normalize column names
    df.columns = [col.strip().replace(" ", "_").title() for col in df.columns]

    # Auto parse dates
    for col in df.columns:
        if "date" in col.lower():
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            except:
                pass

    return df

# -----------------------------
# EXCEL STYLING
# -----------------------------
def style_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    wb.save(file_path)

# -----------------------------
# SUMMARY REPORT
# -----------------------------
def generate_summary(df):
    print(Fore.GREEN + "\n===== DATA SUMMARY =====")
    print(df.describe(include='all'))

    print(Fore.GREEN + "\n===== NULL VALUES =====")
    print(df.isnull().sum())

# -----------------------------
# MAIN
# -----------------------------
def main():
    parser = argparse.ArgumentParser(
        description="CSV → Excel Converter Ultra"
    )

    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Input CSV file"
    )

    parser.add_argument(
        "-o",
        "--output",
        default="output.xlsx",
        help="Output Excel file"
    )

    args = parser.parse_args()

    input_file = args.input
    output_file = args.output

    print(Fore.BLUE + "\n🚀 CSV → Excel Converter Ultra")

    # File exists check
    if not os.path.exists(input_file):
        print(Fore.RED + "[-] Input file not found!")
        logging.error("Input file not found")
        return

    try:
        # Read CSV
        print(Fore.CYAN + "[+] Reading CSV file...")
        df = pd.read_csv(input_file)

        logging.info("CSV loaded successfully")

        # Clean data
        df = clean_data(df)

        # Generate summary
        generate_summary(df)

        # Export Excel
        print(Fore.CYAN + f"\n[+] Exporting to {output_file}...")
        df.to_excel(output_file, index=False)

        # Style Excel
        style_excel(output_file)

        logging.info("Excel exported successfully")

        print(Fore.GREEN + "\n✅ Conversion Completed Successfully!")

    except Exception as e:
        print(Fore.RED + f"\n[-] Error: {e}")
        logging.error(str(e))


if __name__ == "__main__":
    main()
